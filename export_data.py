"""
export_data.py
Lee el Excel de seguimiento (carpeta padre) y genera data/dashboard_data.json.
Ejecutar: python export_data.py
"""

import openpyxl
import json
import os
from datetime import datetime

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(SCRIPT_DIR, "..", "APOYO Consultoría - Seguimiento.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "data", "dashboard_data.json")

def ss(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None

def clean_dist(d):
    if not d: return None
    return str(d).replace("Distrito Judicial: ", "").strip()

def main():
    print(f"Leyendo: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── Seguimiento entrevistas ────────────────────────────────────────────
    # Fila 5 (índice 4) = encabezados, datos desde fila 6 (índice 5)
    ws_seg = wb["Seguimiento entrevistas"]
    seguimiento = []
    for row in list(ws_seg.iter_rows(values_only=True))[5:]:
        if row[2] is None: continue          # sin nombre = vacío
        seguimiento.append({
            "ambito":      clean_dist(row[0]),
            "nombre":      ss(row[2]),
            "institucion": ss(row[3]),
            "categoria":   ss(row[7]),
            "tipo":        ss(row[8]),
            "subtipo":     ss(row[9]),
            "contactado":  ss(row[18]),
            "estado":      ss(row[19]),
            "modalidad":   ss(row[26]),
            "duracion":    row[27],
        })

    # ── Grupos focales ─────────────────────────────────────────────────────
    ws_gf = wb["Grupos focales"]
    grupos = []
    for row in list(ws_gf.iter_rows(values_only=True))[1:]:
        if row[0] is None: continue
        grupos.append({
            "id":          ss(row[0]),
            "tipo_gf":     ss(row[1]),
            "nombre":      ss(row[2]),
            "institucion": ss(row[3]),
            "actor":       ss(row[7]),
            "distrito":    clean_dist(row[8]),
            "encargado":   ss(row[9]),
            "contacto":    ss(row[11]),
            "respuesta":   ss(row[15]),
            "fecha_gf":    ss(row[16]),
            "asistencia":  ss(row[17]),
        })

    # ── Entrevistas completadas ────────────────────────────────────────────
    ws_ent = wb["Entrevistas"]
    entrevistas = []
    for row in list(ws_ent.iter_rows(values_only=True))[1:]:
        if row[0] is None: continue
        entrevistas.append({
            "id":          ss(row[0]),
            "nombre":      ss(row[1]),
            "institucion": ss(row[2]),
            "tipo":        ss(row[6]),
            "actor":       ss(row[7]),
            "distrito":    clean_dist(row[8]),
            "encargado":   ss(row[9]),
            "resultado":   ss(row[17]),
            "duracion":    row[18],
            "modalidad":   ss(row[21]),
        })

    # ── Panel resumen (extrae KPIs calculados) ─────────────────────────────
    ws_panel = wb["Panel de seguimiento>>"]
    panel_rows = list(ws_panel.iter_rows(values_only=True))

    # Fila 37 (índice 36) en adelante tiene el resumen de estados
    estados = {}
    for row in panel_rows[37:46]:
        vals = [v for v in row if v is not None]
        if len(vals) >= 2 and isinstance(vals[1], (int, float)):
            estados[str(vals[0])] = int(vals[1])

    # Fila 38 (índice 37): ámbito vs realizadas/meta
    ambitos = {}
    for row in panel_rows[38:50]:
        vals = [v for v in row if v is not None]
        if len(vals) >= 3 and isinstance(vals[1], (int, float)) and isinstance(vals[2], (int, float)):
            # formato: [ambito_label, realizadas, meta, pct]
            if isinstance(vals[0], str) and vals[0] not in ('Total',):
                ambitos[vals[0]] = {"realizadas": int(vals[1]), "meta": int(vals[2])}

    data = {
        "last_updated": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "seguimiento":  seguimiento,
        "grupos":       grupos,
        "entrevistas":  entrevistas,
        "estados":      estados,
        "ambitos":      ambitos,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ JSON guardado en: {OUTPUT_PATH}")
    print(f"  Seguimiento:  {len(seguimiento)}")
    print(f"  Grupos:       {len(grupos)}")
    print(f"  Entrevistas:  {len(entrevistas)}")

if __name__ == "__main__":
    main()
